import os
import sys

from docxtpl import DocxTemplate
import docx2pdf
from openpyxl import load_workbook
from app.helpers import models_helper
from num2words import num2words


class DocumentMaker:
    __context: dict
    __folder_path: str
    __unique_name: str

    def __init__(self, context: dict):
        self.__context = context

    def make_documents_in_folder(self, path: str) -> str:
        self.__folder_path = path
        self.__make_folder()
        self.__render_docx()
        self.__render_excel()
        return self.__folder_path

    def __make_folder(self):
        self.__unique_name = self.__context['shortname'] + " " + self.__context['pib_buyer'] + " " + self.__context[
            'settle_type'] + " " + self.__context['settle_name']
        self.__folder_path = self.__folder_path + '/' + self.__unique_name
        os.mkdir(self.__make_unique_folder_name())

    def __make_unique_folder_name(self):
        if not os.path.exists(self.__folder_path):
            return self.__folder_path
        i = 1
        while os.path.exists(self.__folder_path + '_' + str(i)):
            i += 1
        self.__folder_path = self.__folder_path + '_' + str(i)
        return self.__folder_path

    def __render_docx(self):
        doc = DocxTemplate("templates/Template.docx")
        self.__set_to_context(doc)
        doc.render(self.__context)
        document_name = self.__folder_path + '/' + "Д+с " + self.__unique_name + ".docx"
        doc.save(document_name)
        if sys.platform in ("darwin", "win32"):
            docx2pdf.convert(document_name)

    def __set_to_context(self, doc):
        self.__context['image'] = models_helper.get_model_image(doc, self.__context['model'])

    def __render_excel(self):
        wb = load_workbook('templates/Template.xlsx')
        ws = wb.active
        number_and_date: str = " № " + self.__context['contract_date'] + '/' + str(
            self.__context['contract_number']) + \
                               " від " + self.__context['contract_date_ua'] + "р."
        ws['B16'] = "Рахунок на оплату" + number_and_date
        ws['I21'] = self.__context['pib_buyer']
        ws['H24'] = number_and_date
        ws['E28'] = self.__context['model']
        ws['E28'] = self.__context['model']
        true_price_10 = str(self.__context['true_price_10_uah'])
        ws['AQ28'] = true_price_10
        ws['AU28'] = true_price_10
        ws['AU30'] = true_price_10
        ws['B33'] = "Всього найменувань 1, на суму " + true_price_10 + " грн."
        price_in_words: str = num2words(self.__context['true_price_10_uah'], lang='uk')
        price_in_words: str = price_in_words.replace(' нуль', '')
        price_in_words: str = price_in_words.replace(' кома', '', 1)
        ws['B34'] = price_in_words + " гривні 00 копійок"
        wb.save(self.__folder_path + '/' + "рахунок " + self.__unique_name + ".xlsx")
